# -*- coding: utf-8 -*-
import os
from datetime import datetime
import numpy as np
import shutil
from flask import Flask, render_template, redirect, url_for, request, session, send_from_directory, Response
from flask_uploads import UploadSet, configure_uploads, IMAGES, patch_request_class
from flask_wtf import FlaskForm
from flask_wtf.file import FileField, FileRequired, FileAllowed
from wtforms import SubmitField
import xlsxwriter
import xlrd
from xlutils.copy import copy
from flask import jsonify
import openpyxl
import json
from util.AHP import AHP
from flask_sqlalchemy import SQLAlchemy
import config
from flask.json import JSONEncoder as _JSONEncoder


class JSONEncoder(_JSONEncoder):
    def default(self, o):
        if isinstance(o, datetime):
            return int(o.timestamp())
        if hasattr(o, 'keys') and hasattr(o, '__getitem__'):
            return dict(o)
        raise None


app = Flask(__name__)
app.json_encoder = JSONEncoder
app.config.from_object(config)
app.config['SECRET_KEY'] = 'I have a dream'
address = 'C:\\Users\\Administrator\\Desktop\\images\\static\\'
app.config['UPLOADED_PHOTOS_DEST'] = address
app.config['MAX_CONTENT_LENGTH'] = 2 * 1024 * 1024
db = SQLAlchemy(app)
photos = UploadSet('photos', IMAGES)
configure_uploads(app, photos)
patch_request_class(app, size=None)  # set maximum file size, default is 16MB


class UploadForm(FlaskForm):
    photo = FileField(validators=[FileAllowed(photos, u'只能是照片格式!'), FileRequired(u'Choose a file!')])
    submit = SubmitField(u'上传')


@app.route('/', methods=['GET', 'POST'])
def index():
    return render_template('index.html')


@app.route('/index', methods=['GET', 'POST'])
def upload_file():
    folder_name = request.form.get('folderName')
    # form = UploadForm()
    folder = address + folder_name
    tasks = Task.query.filter_by(folder_name=folder_name).all()
    if len(tasks) == 0:
        task = Task(folder_name=folder_name, size=len(request.files.getlist('photo')), status='0', create_time=datetime.now())
        # 调用添加方法
        db.session.add(task)
        db.session.commit()
    else:
        task = Task.query.filter_by(folder_name=folder_name).first()
        task.size = str(int(task.size) + len(request.files.getlist('photo')))
        db.session.commit()

    if not os.path.exists(folder):
        os.makedirs(folder)
    full_path = folder + '\\names.txt'
    file = open(full_path, 'a')
    # create_excel(len(request.files.getlist('photo')))
    for filename in request.files.getlist('photo'):
        name = filename.filename
        file.write(name + '\n')
        photos.save(filename, folder=folder, name=name)
    task = Task.query.filter_by(folder_name=folder_name).first()
    return jsonify(task)


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8080, debug=True)
    # app.run(debug=True)


@app.route('/page_list', methods=['GET', 'POST'])
def page_list():
    user_id = request.headers.get('Authorization',None)
    task = Task.query.filter_by(user_id=user_id, status=2).first()
    folder_name = address + task.folder_name
    if not os.path.exists(folder_name):
        return jsonify(0)
    files_list = os.listdir(folder_name)

    return jsonify(len(files_list) - 3)


def create_excel(size, folder_name):
    # 新建一个Excel文件
    wb = openpyxl.Workbook()
    ws1 = wb.active
    for i in range(size - 1):
        ws1.cell(row=i+1, column=i+1, value=1)
    wb.save((folder_name + '\\data.xlsx'))
    workbook = xlsxwriter.Workbook(folder_name + '\\result.xlsx')

    workbook.close()


@app.route('/submit', methods=['GET', 'POST'])
def submit():
    user_id = request.headers.get('Authorization', None)
    task = Task.query.filter_by(user_id=user_id, status=2).first()
    task.status = 3
    db.session.commit()

    folder_name = address + task.folder_name
    filename = folder_name + "\\data.xlsx"
    arr = []
    ex = xlrd.open_workbook(filename).sheets()[0]
    for i in range(ex.nrows):
        col = ex.row_values(i)
        for index, n in enumerate(col):
            if isinstance(n, str):
                col[index] = 0
        arr.append(col)
    M = np.array(arr)
    obj = AHP(M)
    evec = obj.get_evec(obj.supp_mat(M))
    obj.save_result(evec, folder_name)
    return jsonify("success")


@app.route('/update_excel/<row>/<line>/<value>', methods=['GET', 'POST'])
def update_excel(row, line, value):
    user_id = request.headers.get('Authorization', None)
    task = Task.query.filter_by(user_id=user_id, status=2).first()
    folder_name = address + task.folder_name
    row = int(row) - 1
    line = int(line) - 1
    xls = xlrd.open_workbook(folder_name + '\\data.xlsx')
    xlsc = copy(xls)
    shtc = xlsc.get_sheet(0)
    shtc.write(int(row), int(line), int(value))
    xlsc.save(folder_name + '\\data.xlsx')
    return jsonify("success")


@app.route('/open/<filename>', methods=['GET', 'POST'])
def open_file(filename):
    user_id = request.headers.get('Authorization', None)
    task = Task.query.filter_by(user_id=user_id, status=2).first()
    folder_name = address + task.folder_name
    line = getline(folder_name + "\\names.txt", int(filename))
    name = line.replace("\n", "")
    global app
    app.config['UPLOADED_PHOTOS_DEST'] = folder_name

    global photos
    photos = UploadSet('photos', IMAGES)
    configure_uploads(app, photos)

    file_url = photos.url(name)
    return jsonify(file_url)


@app.route('/delete/<filename>')
def delete_file(filename):
    file_path = photos.path(filename)
    os.remove(file_path)
    return render_template('manage.html', files_list=files_list)


@app.route('/download/<folder_name>', methods=['GET', 'POST'])
def download(folder_name):
    folder_name = address + folder_name
    # filename = folder_name + "\\data.xlsx"
    # arr = []
    # ex = xlrd.open_workbook(filename).sheets()[0]
    # for i in range(ex.nrows):
    #     col = ex.row_values(i)
    #     for index, n in enumerate(col):
    #         if isinstance(n, str):
    #             col[index] = 0
    #     arr.append(col)
    # M = np.array(arr)
    # obj = AHP(M)
    # evec = obj.get_evec(obj.supp_mat(M))
    # obj.save_result(evec, folder_name)
    return send_from_directory(folder_name, filename="result.xlsx", as_attachment=True)


def getline(the_file_path, line_number):
    if line_number < 1:
        return ''
    for cur_line_number, line in enumerate(open(the_file_path, 'rU')):
        if cur_line_number == line_number-1:
            return line
    return ''


@app.route('/getValue/<row>/<line>', methods=['GET', 'POST'])
def get_excel(row, line):
    user_id = request.headers.get('Authorization', None)
    task = Task.query.filter_by(user_id=user_id, status=2).first()
    folder_name = address + task.folder_name
    row = int(row) - 1
    line = int(line) - 1
    x1 = xlrd.open_workbook(folder_name + '\\data.xlsx')
    sheet1 = x1.sheet_by_index(0)
    a12 = sheet1.cell_value(row, line)
    return jsonify(a12)


@app.route('/login', methods=['POST'])
def login():
    data = request.get_data()
    json_data = json.loads(data.decode("utf-8"))
    username = json_data.get("username")
    password = json_data.get("password")
    user = User.query.filter_by(username=username, password=password).all()
    if len(user) == 1:
        return jsonify({'status':'ok','info':'%s登录成功'%username,'session':user[0].id,'role':user[0].role})
    return jsonify({'status':'no','info':'登录失败'})


@app.route('/registry', methods=['POST'])
def registry():
    data = request.get_data()
    json_data = json.loads(data.decode("utf-8"))
    username = json_data.get("username")
    password = json_data.get("password")
    users = User.query.filter_by(username=username).all()
    if len(users) > 0:
        return jsonify({'status':'no','info':'%s注册失败'%username})
    else:
        user = User(username=username, password=password)
        # 调用添加方法
        db.session.add(user)
        db.session.commit()
        return jsonify({'status':'ok','info':'%s注册成功'%username,'session':username,'role':users[0].role})


@app.route('/getTask', methods=['GET'])
def get_task():
    tasks = Task.query.order_by(Task.create_time.desc()).all()
    return jsonify(tasks)


@app.route('/getUsers', methods=['GET'])
def get_users():
    users = User.query.all()
    return jsonify(users)


@app.route('/deleteTask/<task_id>', methods=['GET'])
def delete_task(task_id):
    task = Task.query.filter_by(id=task_id).first()
    folder_name = address + task.folder_name
    shutil.rmtree(path=folder_name)

    Task.query.filter_by(id=task_id).delete()
    db.session.commit()
    return jsonify('success')


@app.route('/updateTask', methods=['POST'])
def update_task():
    data = request.get_data()
    json_data = json.loads(data.decode("utf-8"))
    task_id = json_data.get("id")
    user_id = json_data.get("user_id")
    status = json_data.get("status")
    folder_name = json_data.get("folder_name")
    if int(status) == 2:
        files_list = os.listdir(address + str(folder_name))
        create_excel(len(files_list), address + str(folder_name))
    task = Task.query.filter_by(id=task_id).first()
    task.user_id = user_id
    task.status = status
    db.session.commit()

    # user_id = request.headers.get('Authorization',None)
    users = User.query.all()
    return jsonify(users)


class User(db.Model):
    __tablename__ = 'user'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    username = db.Column(db.String(100), nullable=False)
    password = db.Column(db.String(100), nullable=False)
    role = db.Column(db.String(100), nullable=False)

    def keys(self):
        return ['id', 'username', 'password', 'role']

    def __getitem__(self, item):
        return getattr(self, item)


class Task(db.Model):
    __tablename__ = 'task'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    user_id = db.Column(db.String(100), nullable=False)
    folder_name = db.Column(db.String(100), nullable=False)
    status = db.Column(db.String(100), nullable=False)
    size = db.Column(db.String(100), nullable=False)
    create_time = db.Column(db.DateTime, nullable=False)  # 发送时间

    def keys(self):
        return ['id', 'user_id', 'folder_name', 'status', 'size', 'create_time']

    def __getitem__(self, item):
        return getattr(self, item)



